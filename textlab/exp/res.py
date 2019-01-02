#! python3
# -*- coding:utf-8 -*-
'''
隐藏结果, 统计, 成功率, 容量
'''
import os
import numpy as np
import matplotlib.pyplot as plt
from gensim.models import Word2Vec
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from fileutil import FileUtil
from word2vec.vector import WV


class Result(object):

    def __init__(self, model):
        self.model = model

    def res(self, filename):
        SR = []
        HC = []

        with open(filename, 'r+', encoding='utf-8') as fin:
            for line in fin.readlines():
                pre, post = line.split('||')
                pre = pre.strip().split('\t')
                post = post.strip().split('\t')
                sr = int(pre[2])/int(pre[1])
                SR.append(sr)

                for p in post:
                    hc = int(p)
                    if hc > 0 and hc < 100:
                        HC.append(hc)
                pass
        # zip(SR,HC)
        # x1 = np.linspace(0, len(SR),len(SR))
        # x2 = np.linspace(0, len(HC), len(HC))
        #
        # plt.subplot(2, 1, 1)
        # plt.plot(x1, SR)
        #
        # plt.subplot(2, 1, 2)
        # plt.plot(x2, HC)
        # plt.show()
        return SR, HC

    def CR(self, origin_info, extract_info):
        return WV.wmd(self.model, origin_info, extract_info)

    def all_CR(self, infopath, extpath):
        CR = []
        file_list = os.listdir(infopath)
        for name in file_list:
            origin_info = FileUtil.readfile(filename=os.path.join(infopath, name))
            ext_info = FileUtil.readfile(filename=os.path.join(extpath, name))
            cr = self.CR(origin_info, ext_info)
            CR.append(cr)
        return CR

    # 写入 excel
    @staticmethod
    def save(filename, data):
        wb = load_workbook(filename=filename)
        ws = wb['data']
        # wb = Workbook()
        # ws = wb.create_sheet(title='data')
        col = ws.max_column
        for j, item in enumerate(data):
            ws['{0}{1}'.format(get_column_letter(col + 1), str(j + 1))] = data[j]
        wb.save(filename='col.xlsx')


if __name__ == '__main__':
    # models = Word2Vec.load('./m.bin')
    # r = Result(model=models)
    # r.res(r'F:\LabData\NetBigData\test\res.txt')
    a = [1,2,3,4,5,6,7]
